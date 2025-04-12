
from flask import Flask, render_template, request, redirect, url_for
import sqlite3
from openpyxl import load_workbook
from io import BytesIO
from flask import session

app = Flask(__name__)
app.secret_key = "kanku_secret_2024"  # можеш замінити на свій

DATABASE = "kanku_kata.db"

def get_db():
    conn = sqlite3.connect(DATABASE)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_db()
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS athletes (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    name TEXT NOT NULL,
                    category TEXT NOT NULL,
                    coach TEXT NOT NULL
                )''')
    c.execute('''CREATE TABLE IF NOT EXISTS scores (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    athlete_id INTEGER,
                    s1 REAL, s2 REAL, s3 REAL, s4 REAL, s5 REAL,
                    average REAL,
                    FOREIGN KEY (athlete_id) REFERENCES athletes(id)
                )''')
    conn.commit()
    conn.close()

@app.route("/")
def index():
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM athletes")
    athletes = c.fetchall()
    grouped = {}
    for a in athletes:
        grouped.setdefault(a["category"], []).append(a)
    return render_template("index.html", grouped=grouped)

@app.route("/add_athlete", methods=["POST"])
def add_athlete():
    name = request.form["name"]
    category = request.form["category"]
    coach = request.form["coach"]
    conn = get_db()
    c = conn.cursor()
    c.execute("INSERT INTO athletes (name, category, coach) VALUES (?, ?, ?)", (name, category, coach))
    conn.commit()
    return redirect(url_for("index"))

@app.route("/score/<int:athlete_id>", methods=["GET", "POST"])
def score(athlete_id):
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM athletes WHERE id=?", (athlete_id,))
    athlete = c.fetchone()

    if request.method == "POST":
        s1 = float(request.form["s1"])
        s2 = float(request.form["s2"])
        s3 = float(request.form["s3"])
        s4 = float(request.form["s4"])
        s5 = float(request.form["s5"])
        scores = sorted([s1, s2, s3, s4, s5])
        average = round(sum(scores[1:4]) / 3, 2)

        c.execute("INSERT INTO scores (athlete_id, s1, s2, s3, s4, s5, average) VALUES (?, ?, ?, ?, ?, ?, ?)",
          (athlete_id, s1, s2, s3, s4, s5, average))
        conn.commit()
        return redirect(url_for("index"))

    return render_template("score.html", athlete=athlete)

@app.route("/import_excel", methods=["POST"])
def import_excel():
    file = request.files.get("file")
    if not file:
        return "Файл не завантажено", 400

    wb = load_workbook(filename=BytesIO(file.read()))
    ws = wb.active

    group = []
    conn = get_db()
    c = conn.cursor()

    def process_group(rows):
        if not rows:
            return
        years = [int(r[1]) for r in rows if r[1]]
        belts = [int(r[2]) for r in rows if r[2]]
        gender = rows[0][3]
        min_year, max_year = min(years), max(years)
        min_belt, max_belt = min(belts), max(belts)
        category_name = f"{min_year}-{max_year} / {min_belt}-{max_belt} кю / {gender}"

        for r in rows:
            name = r[0]
            coach = r[4]
            tatami = r[5]
            if name and coach:
                c.execute(
                    "INSERT INTO athletes (name, category, coach, tatami) VALUES (?, ?, ?, ?)",
                    (name, category_name, coach, tatami)
                )

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row[:4]):
            process_group(group)
            group = []
        else:
            group.append(row)

    process_group(group)
    conn.commit()
    conn.close()
    return redirect(url_for("index"))

@app.route("/results")
def results():
    conn = get_db()
    c = conn.cursor()
    c.execute("""
        SELECT a.name, a.category, a.coach,
               s.s1, s2, s3, s4, s5, s.average
        FROM athletes a
        JOIN scores s ON a.id = s.athlete_id
    """)
    rows = c.fetchall()

    grouped = {}
    for row in rows:
        category = row["category"]
        grouped.setdefault(category, []).append({
            "name": row["name"],
            "coach": row["coach"],
            "scores": [row["s1"], row["s2"], row["s3"], row["s4"], row["s5"]],
            "average": row["average"]
        })

    return render_template("results.html", results=grouped)

@app.route("/bracket", methods=["GET", "POST"])
def bracket():
    conn = get_db()
    c = conn.cursor()

    # Всі татамі
    c.execute("SELECT DISTINCT tatami FROM athletes ORDER BY tatami")
    all_tatamis = [row["tatami"] for row in c.fetchall()]
    selected_tatami = request.args.get("tatami") or request.form.get("tatami") or (all_tatamis[0] if all_tatamis else None)

    # Категорії на обраному татамі
    c.execute("SELECT DISTINCT category FROM athletes WHERE tatami=? ORDER BY category", (selected_tatami,))
    all_categories = [row["category"] for row in c.fetchall()]
    selected_category = request.args.get("category") or request.form.get("category") or (all_categories[0] if all_categories else None)

    # Збереження переможця
    if request.method == "POST" and "winner" in request.form:
        winner_id = int(request.form["winner"])
        category = request.form["category"]
        tatami = request.form["tatami"]
        current_round = int(request.form["round"])
        c.execute("INSERT OR IGNORE INTO bracket_results (athlete_id, category, round) VALUES (?, ?, ?)", (winner_id, category, current_round))
        conn.commit()
        return redirect(url_for("bracket", category=category, tatami=tatami))

    # Завантаження спортсменів з оцінками
    c.execute("""
        SELECT a.id, a.name, s.average
        FROM athletes a
        JOIN scores s ON a.id = s.athlete_id
        WHERE a.category=? AND a.tatami=?
        ORDER BY s.average DESC
    """, (selected_category, selected_tatami))
    base_athletes = [{"id": row["id"], "name": row["name"], "average": row["average"]} for row in c.fetchall()]

    # Поточний максимум раунду
    c.execute("SELECT MAX(round) FROM bracket_results WHERE category=?", (selected_category,))
    max_round = c.fetchone()[0] or 1

    def generate_seeding(athletes):
        count = len(athletes)
        order = {
            8: [0,7,3,4,2,5,1,6],
            7: [0,3,2,4,1,5,6],
            6: [0,3,2,4,1,5],
            5: [0,3,2,1,4],
            4: [0,3,2,1],
            3: [0,2,1],
            2: [0,1],
            1: [0],
        }.get(count, list(range(count)))
        return [athletes[i] for i in order if i < count]

    def get_athletes_for_round(round_num):
        if round_num == 1:
            return generate_seeding(base_athletes)
        c.execute("""
            SELECT a.id, a.name FROM athletes a
            JOIN bracket_results br ON a.id = br.athlete_id
            WHERE br.category=? AND br.round=?
        """, (selected_category, round_num - 1))
        return [{"id": row["id"], "name": row["name"]} for row in c.fetchall()]

    def build_bracket_for_round(round_num):
        athletes = get_athletes_for_round(round_num)
        bracket = []
        for i in range(0, len(athletes), 2):
            p1 = athletes[i]
            p2 = athletes[i+1] if i+1 < len(athletes) else {"id": None, "name": "(бай)"}
            winner = None
            if p1["id"] and p2["id"]:
                c.execute("SELECT athlete_id FROM bracket_results WHERE category=? AND round=? AND athlete_id IN (?, ?)", (selected_category, round_num, p1["id"], p2["id"]))
                row = c.fetchone()
                if row:
                    winner = p1["name"] if row["athlete_id"] == p1["id"] else p2["name"]
            bracket.append({
                "player1": p1["name"], "player2": p2["name"],
                "player1_id": p1["id"], "player2_id": p2["id"],
                "winner": winner,
                "round": round_num
            })
        return bracket

    all_brackets = []
    for r in range(1, max_round + 2):
        matches = build_bracket_for_round(r)
        if matches:
            all_brackets.append({"round": r, "matches": matches})
            if len(matches) == 1 and (matches[0]["player2"] == "(бай)" or matches[0]["winner"]):
                break

    return render_template("bracket.html",
                           all_brackets=all_brackets,
                           selected_category=selected_category,
                           all_categories=all_categories,
                           selected_tatami=selected_tatami,
                           all_tatamis=all_tatamis)


@app.route("/bracket/round", methods=["GET", "POST"])
def bracket_round():
    if "bracket_state" not in session:
        conn = get_db()
        c = conn.cursor()
        c.execute("SELECT a.name, a.category, s.average FROM athletes a JOIN scores s ON a.id = s.athlete_id")
        rows = c.fetchall()
        grouped = {}
        for r in rows:
            grouped.setdefault(r["category"], []).append({"name": r["name"], "average": r["average"]})
        session["bracket_state"] = grouped
        session["rounds"] = {}

    if request.method == "POST":
        round_id = request.form["round"]
        rounds = session["rounds"]
        prev_matches = session.get("current_matches", [])
        winners = []
        for i, match in enumerate(prev_matches):
            selected = int(request.form.get(f"winner_{i}", 0))
            winners.append(match[selected])
        rounds[round_id] = winners
        session["rounds"] = rounds
        session.modified = True

    current_round = len(session["rounds"]) + 1
    current_matches = []
    if current_round == 1:
        current_matches = []
        for category, players in session["bracket_state"].items():
            current_matches += pair_players(players)
    else:
        all_winners = []
        for cat_round in session["rounds"].values():
            all_winners += cat_round
        current_matches = pair_players(all_winners)

    session["current_matches"] = current_matches
    return render_template("bracket_round.html",
                           matches=current_matches,
                           round_id=f"round_{current_round}",
                           round_name=f"{current_round}")

@app.route("/winners")
def winners():
    if "rounds" not in session:
        return render_template("winners.html", winners=[])

    rounds = session["rounds"]
    round_keys = sorted(rounds.keys())
    final_winners = []

    if len(round_keys) >= 2:
        last_round = rounds[round_keys[-1]]
        second_last_round = rounds[round_keys[-2]]

        if len(last_round) >= 1:
            final_winners.append(last_round[0])
        if len(last_round) >= 2:
            final_winners.insert(1, last_round[1])
        if len(second_last_round) >= 2:
            third_place = second_last_round[1] if second_last_round[0] in last_round else second_last_round[0]
            fourth_place = second_last_round[0] if second_last_round[0] in last_round else second_last_round[1]
            final_winners.append(third_place)
            final_winners.append(fourth_place)

    return render_template("winners.html", winners=final_winners)

@app.route("/clear", methods=["GET", "POST"])
def clear():
    if request.method == "POST":
        session.pop("bracket_state", None)
        session.pop("rounds", None)
        session.pop("current_matches", None)
        return redirect(url_for("index"))
    return render_template("confirm_clear.html")

@app.route("/evaluate", methods=["GET", "POST"])
def evaluate():
    conn = get_db()
    c = conn.cursor()

    selected_category = request.args.get("category") or request.form.get("category")
    selected_tatami = request.args.get("tatami") or request.form.get("tatami")
    selected_athlete_id = request.args.get("athlete_id")
    only_unscored = request.args.get("only_unscored") == "1" or request.form.get("only_unscored") == "1"

    if request.method == "POST":
        athlete_id = int(request.form["athlete_id"])
        scores = [float(request.form[f"s{i}"]) for i in range(1, 6)]
        sorted_scores = sorted(scores)
        avg = round(sum(sorted_scores[1:4]) / 3, 2)
        edit_mode = request.form.get("edit_mode") == "1"

        if edit_mode:
            c.execute("UPDATE scores SET s1=?, s2=?, s3=?, s4=?, s5=?, average=? WHERE athlete_id=?",
                      (*scores, avg, athlete_id))
        else:
            c.execute("INSERT INTO scores (athlete_id, s1, s2, s3, s4, s5, average) VALUES (?, ?, ?, ?, ?, ?, ?)",
                      (athlete_id, *scores, avg))
        conn.commit()

        c.execute("""
            SELECT a.id FROM athletes a
            WHERE a.category=? AND a.tatami=? AND a.id NOT IN (SELECT athlete_id FROM scores)
            ORDER BY a.id
        """, (selected_category, selected_tatami))
        next_athlete = c.fetchone()
        if next_athlete:
            return redirect(url_for("evaluate", tatami=selected_tatami, category=selected_category,
                                    athlete_id=next_athlete["id"], only_unscored=int(only_unscored)))
        else:
            return redirect(url_for("evaluate", tatami=selected_tatami, category=selected_category,
                                    only_unscored=int(only_unscored)))

    if not selected_tatami:
        c.execute("SELECT DISTINCT tatami FROM athletes ORDER BY tatami")
        tatami_list = [row["tatami"] for row in c.fetchall()]
        return render_template("evaluate.html", categories=[], tatami_list=tatami_list,
                               selected_category=None, selected_tatami=None, only_unscored=False)

    # Підрахунок категорій з оціненими
    c.execute("SELECT DISTINCT category FROM athletes WHERE tatami=? ORDER BY category", (selected_tatami,))
    raw_categories = [row["category"] for row in c.fetchall()]
    categories = []

    for cat in raw_categories:
        c.execute("SELECT COUNT(*) FROM athletes WHERE category=? AND tatami=?", (cat, selected_tatami))
        total = c.fetchone()[0]
        c.execute("SELECT COUNT(*) FROM scores WHERE athlete_id IN (SELECT id FROM athletes WHERE category=? AND tatami=?)", (cat, selected_tatami))
        scored = c.fetchone()[0]
        unscored = total - scored
        categories.append({"label": f"{cat} — ✅ {scored}/{total} ⏳ {unscored}", "value": cat})

    if not selected_category:
        c.execute("SELECT DISTINCT tatami FROM athletes ORDER BY tatami")
        tatami_list = [row["tatami"] for row in c.fetchall()]
        return render_template("evaluate.html", categories=categories, tatami_list=tatami_list,
                               selected_category=None, selected_tatami=selected_tatami, only_unscored=only_unscored)

    c.execute("SELECT a.id, a.name, a.coach FROM athletes a WHERE a.category=? AND a.tatami=? ORDER BY a.id",
              (selected_category, selected_tatami))
    all_athletes = c.fetchall()
    athlete_list = []
    unscored_count = 0
    for a in all_athletes:
        c.execute("SELECT average FROM scores WHERE athlete_id=?", (a["id"],))
        score_data = c.fetchone()
        scored = score_data is not None
        avg = score_data["average"] if scored else None
        if only_unscored and scored:
            continue
        if not scored:
            unscored_count += 1
        athlete_list.append({
            "id": a["id"],
            "name": a["name"],
            "coach": a["coach"],
            "scored": scored,
            "average": avg
        })

    total = len(all_athletes)

    athlete = None
    athlete_scores = None
    previous_id = None
    next_id = None
    if selected_athlete_id:
        selected_athlete_id = int(selected_athlete_id)
        c.execute("SELECT a.id, a.name, a.coach FROM athletes a WHERE a.id=?", (selected_athlete_id,))
        athlete = c.fetchone()
        c.execute("SELECT s1, s2, s3, s4, s5 FROM scores WHERE athlete_id=?", (athlete["id"],))
        row = c.fetchone()
        if row:
            athlete_scores = [row["s1"], row["s2"], row["s3"], row["s4"], row["s5"]]

        ids = [a["id"] for a in athlete_list]
        if selected_athlete_id in ids:
            index = ids.index(selected_athlete_id)
            previous_id = ids[index - 1] if index > 0 else None
            next_id = ids[index + 1] if index < len(ids) - 1 else None

    c.execute("SELECT DISTINCT tatami FROM athletes ORDER BY tatami")
    tatami_list = [row["tatami"] for row in c.fetchall()]

    return render_template("evaluate.html",
                           selected_category=selected_category,
                           selected_tatami=selected_tatami,
                           athlete=athlete,
                           athlete_scores=athlete_scores,
                           athlete_list=athlete_list,
                           total=total,
                           unscored_count=unscored_count,
                           categories=categories,
                           tatami_list=tatami_list,
                           only_unscored=only_unscored,
                           previous_id=previous_id,
                           next_id=next_id)

@app.route("/clear_categories", methods=["GET", "POST"])
def clear_categories():
    if request.method == "POST":
        conn = get_db()
        c = conn.cursor()
        # Спочатку видаляємо оцінки (залежність від athlete_id)
        c.execute("DELETE FROM scores")
        # Потім видаляємо спортсменів (разом із категоріями)
        c.execute("DELETE FROM athletes")
        conn.commit()
        return redirect(url_for("index"))
    
    return '''
        <h2>Очистити всі категорії і спортсменів?</h2>
        <form method="post">
            <button type="submit" style="padding:10px 20px; background:red; color:white; border:none; border-radius:5px;">Очистити</button>
            <a href="/" style="margin-left:15px;">Скасувати</a>
        </form>
    '''

if __name__ == "__main__":
    init_db()
    app.run(debug=True)